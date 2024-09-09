import pandas as pd
import time
import os
from openpyxl import load_workbook

# Define the paths to your input and output Excel files
MDG = "C:/Users/2095421/Downloads/Migration/PrakharTest/M1P MARD 1.xlsx"
DataValidationPath = "C:/Users/2095421/Downloads/Migration/PrakharTest/abc.xlsx"

# Record the start time
start_time = time.time()

# Read the large DataFrame from the input Excel file
df1 = pd.read_excel(MDG, sheet_name="Sheet1")

# Add a new column to the DataFrame
df1.insert(0, 'KEY_' + os.path.basename(MDG), df1['MATNR'].astype(str) + df1['WERKS'].astype(str) + df1['LGORT'].astype(str))

# Set the new column as the index of the DataFrame
df1.set_index('KEY_' + os.path.basename(MDG), inplace=True)

# Define the number of rows you want in each chunk
chunk_size = 10000  # Adjust this value based on your needs

# Calculate the number of chunks
num_chunks = len(df1) // chunk_size + 1

# Check if the file exists
if os.path.exists(DataValidationPath):
    book = load_workbook(DataValidationPath)
else:
    book = None

# Create a new Excel writer object
with pd.ExcelWriter(DataValidationPath, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    writer.book = book

    # Loop through the chunks
    for i in range(num_chunks):
        # Calculate the start and end indices for this chunk
        start = i * chunk_size
        end = (i + 1) * chunk_size

        # Get the chunk from the DataFrame
        chunk = df1.iloc[start:end]

        # Check if the sheet exists
        sheet_name = os.path.basename(MDG)
        if sheet_name in writer.sheets:
            startrow = writer.sheets[sheet_name].max_row
        else:
            startrow = 0

        # Write the chunk to the same sheet in the Excel file
        chunk.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)

# Save the Excel file
writer.save()

# Calculate the total memory usage of the DataFrame
total_memory_usage = df1.memory_usage(index=True).sum()
total_memory_usage_mb = total_memory_usage / 1024 / 1024
print(total_memory_usage_mb)

# Record the end time
end_time = time.time()

# Calculate and print the execution time
execution_time = end_time - start_time
print(f"The program took {execution_time} seconds to execute.")
