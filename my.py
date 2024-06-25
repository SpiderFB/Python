import openpyxl

wb = openpyxl.Workbook()
location = 'C:/Users/2095421/Downloads/Migration/'
new_file = 'hello.xlsx'
wb.save(location + new_file)

wb = load_workbook(location + new_file)


MBEW_MDG = "C:/Users/2095421/Downloads/Migration/MBEW_M1Q_Ext.XLSX"
MBEW_ATLAS = "C:/Users/2095421/Downloads/Migration/MBEW_APB_Ext.XLSX"

wb1 = xl.load_workbook(MBEW_MDG)
ws1 = wb1.worksheets['Sheet1']

wb = wb.copy



# wb1 = xl.load_workbook(filename=MBEW_MDG)
# ws1 = wb1.worksheets[0]

# wb2 = xl.load_workbook(filename=MBEW_ATLAS)
# ws2 = wb2.create_sheet(ws1.title)
