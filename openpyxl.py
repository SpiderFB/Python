import openpyxl
from openpyxl import load_workbook

wb = openpyxl.Workbook()
location = 'C:/Users/2095421/Downloads/Migration/'
new_file = 'hello.xlsx'
wb.save(location + new_file)

MBEW_MDG = "C:/Users/2095421/Downloads/Migration/MBEW_M1Q_Ext.XLSX"
MBEW_ATLAS = "C:/Users/2095421/Downloads/Migration/MBEW_APB_Ext.XLSX"

source_wb = load_workbook(MBEW_MDG)
dest_wb = load_workbook(location + new_file)

source_sheet = source_wb['Sheet1']
# dest_sheet = dest_wb.create_sheet('Sheet1_Copy') #if I want to create a new sheet and then copy

dest_sheet = dest_wb['Sheet']

# for row in source_sheet.iter_rows(values_only=True):
#     dest_sheet.append(row)

dest_sheet = source_wb.copy_worksheet(source_sheet)
# dest_sheet.title = 'MBEW_MDG_Copy' #to rename the Sheet in Destination file

dest_wb.save(location+new_file)

source_wb.close()
dest_wb.close()